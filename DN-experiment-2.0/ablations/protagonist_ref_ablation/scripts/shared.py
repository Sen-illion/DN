from __future__ import annotations

import copy
import importlib.util
import json
import math
import os
import shutil
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path
from statistics import mean
from typing import Any, Dict, Iterable, List, Optional

try:
    from openpyxl import Workbook  # type: ignore
    from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore
    from openpyxl.utils import get_column_letter  # type: ignore
except ModuleNotFoundError:  # pragma: no cover
    Workbook = None  # type: ignore
    Alignment = None  # type: ignore
    Font = None  # type: ignore
    PatternFill = None  # type: ignore
    get_column_letter = None  # type: ignore

REPO_ROOT = Path(__file__).resolve().parents[4]
DN_EXPERIMENT_ROOT = REPO_ROOT / "DN-experiment-2.0"
ABLATION_ROOT = DN_EXPERIMENT_ROOT / "ablations" / "protagonist_ref_ablation"
CONFIG_DIR = ABLATION_ROOT / "configs"
DATASETS_DIR = ABLATION_ROOT / "datasets"
RESULTS_DIR = ABLATION_ROOT / "results"
THEMES_JSON = REPO_ROOT / "game_themes_100.json"
LEGACY_ROOT = DN_EXPERIMENT_ROOT / "图片一致性_experiment" / "multiview_image_consistency"
LEGACY_SCORE_SCRIPT = LEGACY_ROOT / "scripts" / "score_image_consistency_per_game.py"
LEGACY_AGGREGATE_SCRIPT = LEGACY_ROOT / "scripts" / "aggregate_multiview_results.py"
EXPORT_IMAGE_PATHS_SCRIPT = DN_EXPERIMENT_ROOT / "export_image_paths_manifest.py"
DEFAULT_CONFIG_PATH = CONFIG_DIR / "protagonist_ref_ablation_config.json"

DEFAULT_GROUP_SPECS = [
    {
        "group_id": "protagonist_ref_0",
        "label": "No protagonist reference",
        "expected_protagonist_ref_count": 0,
    },
    {
        "group_id": "protagonist_ref_1",
        "label": "Single front protagonist reference",
        "expected_protagonist_ref_count": 1,
    },
    {
        "group_id": "protagonist_ref_3",
        "label": "Front/side/back protagonist references",
        "expected_protagonist_ref_count": 3,
    },
]

DEFAULT_DIMENSIONS = [
    "semantic_consistency",
    "subject_attribute_consistency",
    "spatial_consistency",
    "style_lighting_consistency",
    "detail_integrity",
]

_DATASET_KEYS_TO_DROP = {
    "_visual_context",
    "_scene_prompt_cache",
    "_last_scene_prompt_json",
    "_plot_supporting_characters",
}

_SCORE_MODULE: Any = None


def utc_now() -> datetime:
    return datetime.now(timezone.utc)


def utc_timestamp() -> str:
    return utc_now().strftime("%Y%m%d_%H%M%S")


def read_json(path: Path) -> Any:
    for encoding in ("utf-8-sig", "utf-8"):
        try:
            return json.loads(path.read_text(encoding=encoding))
        except (UnicodeDecodeError, json.JSONDecodeError):
            continue
    return json.loads(path.read_text(encoding="utf-8", errors="replace"))


def read_jsonl(path: Path) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    if not path.is_file():
        return rows
    for line in path.read_text(encoding="utf-8").splitlines():
        raw = line.strip()
        if not raw:
            continue
        rows.append(json.loads(raw))
    return rows


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def write_jsonl(path: Path, rows: Iterable[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")


def copy_latest(src: Path, dst: Path) -> None:
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(src, dst)


def json_safe_copy(payload: Any) -> Any:
    return json.loads(json.dumps(payload, ensure_ascii=False, default=str))


def strip_state_for_dataset(global_state: Dict[str, Any]) -> Dict[str, Any]:
    cloned = copy.deepcopy(global_state)
    for key in list(cloned.keys()):
        if key in _DATASET_KEYS_TO_DROP:
            cloned.pop(key, None)
    return json_safe_copy(cloned)


def load_config(path: Path = DEFAULT_CONFIG_PATH) -> Dict[str, Any]:
    if not path.is_file():
        return {}
    return read_json(path)


def load_theme_catalog(path: Path = THEMES_JSON) -> Dict[int, Dict[str, Any]]:
    payload = read_json(path)
    items = payload.get("items") or []
    out: Dict[int, Dict[str, Any]] = {}
    for item in items:
        if isinstance(item, dict) and isinstance(item.get("id"), int):
            out[int(item["id"])] = item
    return out


def resolve_size_settings(config: Dict[str, Any], size: str, overrides: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    presets = config.get("dataset_presets") or {}
    preset = dict(presets.get(size, {}))
    result = {
        "size": size,
        "seed": int(config.get("seed", 20260425)),
        "max_themes": int(preset.get("max_themes", 0) or 0),
        "segments_per_game": int(preset.get("segments_per_game", 0) or 0),
        "difficulty": str((config.get("generation") or {}).get("difficulty", "中等")),
        "tone_key": str((config.get("generation") or {}).get("tone_key", "normal_ending")),
        "start_option": str((config.get("generation") or {}).get("start_option", "开始游戏")),
    }
    if overrides:
        for key, value in overrides.items():
            if value is not None:
                result[key] = value
    return result


def get_group_specs(config: Dict[str, Any]) -> List[Dict[str, Any]]:
    groups = config.get("groups") or DEFAULT_GROUP_SPECS
    out: List[Dict[str, Any]] = []
    for group in groups:
        if not isinstance(group, dict):
            continue
        out.append(
            {
                "group_id": str(group.get("group_id") or "").strip(),
                "label": str(group.get("label") or group.get("group_id") or "").strip(),
                "expected_protagonist_ref_count": int(group.get("expected_protagonist_ref_count", 0) or 0),
            }
        )
    return [group for group in out if group["group_id"]]


def flatten_dict(payload: Dict[str, Any], prefix: str = "") -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for key, value in sorted(payload.items(), key=lambda item: str(item[0])):
        full_key = f"{prefix}.{key}" if prefix else str(key)
        if isinstance(value, dict):
            rows.extend(flatten_dict(value, full_key))
        else:
            rows.append({"key": full_key, "value": json.dumps(value, ensure_ascii=False) if isinstance(value, (list, dict)) else value})
    return rows


def mean_or_none(values: Iterable[Any]) -> Optional[float]:
    numbers = [float(v) for v in values if v is not None]
    return mean(numbers) if numbers else None


def round_or_none(value: Optional[float], digits: int = 4) -> Optional[float]:
    if value is None:
        return None
    return round(float(value), digits)


def safe_text(value: Any, max_len: int = 400) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    text = " ".join(text.split())
    return text[:max_len]


def python_executable() -> str:
    if sys.executable:
        return sys.executable
    candidate = shutil.which("python")
    if candidate:
        return candidate
    raise FileNotFoundError("Unable to locate a Python executable.")


def run_python_script(script_path: Path, args: List[str], cwd: Optional[Path] = None) -> None:
    command = [python_executable(), str(script_path), *args]
    subprocess.run(command, check=True, cwd=str(cwd or REPO_ROOT))


def load_score_module() -> Any:
    global _SCORE_MODULE
    if _SCORE_MODULE is not None:
        return _SCORE_MODULE
    spec = importlib.util.spec_from_file_location("legacy_multiview_score", LEGACY_SCORE_SCRIPT)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Unable to load scorer from {LEGACY_SCORE_SCRIPT}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    _SCORE_MODULE = module
    return module


def get_dimensions() -> List[str]:
    try:
        module = load_score_module()
        dims = list(getattr(module, "DIMENSIONS", []))
        return dims or list(DEFAULT_DIMENSIONS)
    except Exception:
        return list(DEFAULT_DIMENSIONS)


def ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def _normalize_rows(rows: Any) -> List[Dict[str, Any]]:
    if rows is None:
        return []
    if isinstance(rows, list):
        if not rows:
            return []
        if isinstance(rows[0], dict):
            return [dict(item) for item in rows]
        return [{"value": item} for item in rows]
    if isinstance(rows, dict):
        return flatten_dict(rows)
    return [{"value": rows}]


def write_workbook(path: Path, sheet_rows: Dict[str, Any]) -> None:
    if Workbook is None:
        raise RuntimeError("openpyxl is not installed; cannot write xlsx. Re-run with --no-xlsx.")
    ensure_parent(path)
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    wrap_alignment = Alignment(vertical="top", wrap_text=True)

    for sheet_name, raw_rows in sheet_rows.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        rows = _normalize_rows(raw_rows)
        if not rows:
            ws.append(["note"])
            ws.append(["empty"])
            continue
        headers: List[str] = []
        for row in rows:
            for key in row.keys():
                if key not in headers:
                    headers.append(str(key))
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = wrap_alignment
        for row in rows:
            values = []
            for header in headers:
                value = row.get(header, "")
                if isinstance(value, (list, dict)):
                    value = json.dumps(value, ensure_ascii=False)
                values.append(value)
            ws.append(values)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for column_index, header in enumerate(headers, start=1):
            max_len = len(str(header))
            for row_index in range(2, ws.max_row + 1):
                value = ws.cell(row=row_index, column=column_index).value
                max_len = max(max_len, len(str(value)) if value is not None else 0)
                ws.cell(row=row_index, column=column_index).alignment = wrap_alignment
            ws.column_dimensions[get_column_letter(column_index)].width = min(max(max_len + 2, 12), 48)
    wb.save(path)


def latest_dataset_manifest_path() -> Optional[Path]:
    latest = DATASETS_DIR / "latest_dataset_manifest.jsonl"
    return latest if latest.is_file() else None


def slugify(value: str) -> str:
    text = (value or "").strip().lower()
    text = "".join(ch if ch.isalnum() else "_" for ch in text)
    while "__" in text:
        text = text.replace("__", "_")
    return text.strip("_") or "run"


def coverage_ratio(numerator: int, denominator: int) -> float:
    if denominator <= 0:
        return 0.0
    return numerator / denominator


def population_stddev(values: Iterable[Any]) -> Optional[float]:
    numbers = [float(v) for v in values if v is not None]
    if not numbers:
        return None
    if len(numbers) == 1:
        return 0.0
    avg = mean(numbers)
    return math.sqrt(sum((value - avg) ** 2 for value in numbers) / len(numbers))
