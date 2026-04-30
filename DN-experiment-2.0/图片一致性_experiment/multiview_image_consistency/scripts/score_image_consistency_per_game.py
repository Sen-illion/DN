from __future__ import annotations

import argparse
import base64
import json
import os
import re
import sys
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from statistics import mean
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openai import OpenAI
try:
    from openpyxl import Workbook  # type: ignore
except ModuleNotFoundError:  # pragma: no cover
    Workbook = None  # type: ignore

try:
    from dotenv import load_dotenv
except ImportError:
    load_dotenv = None  # type: ignore


REPO_ROOT = Path(__file__).resolve().parents[4]
DEFAULT_EXPERIMENT_ROOT = REPO_ROOT / "DN-experiment-2.0"
DEFAULT_OUTPUT_DIR = REPO_ROOT / "DN-experiment-2.0" / "experiments" / "multiview_image_consistency" / "results"

DIMENSIONS = [
    "semantic_consistency",
    "subject_attribute_consistency",
    "spatial_consistency",
    "style_lighting_consistency",
    "detail_integrity",
]

SCORE_MIN = 1.0
SCORE_MAX = 10.0
SCORE_DEFAULT = 6.0

SYSTEM_PROMPT = """You are a senior visual consistency evaluator.
Evaluate the current image for per-game consistency using available context.

Scoring anchors for each dimension (integer only):
- 10: strongly consistent
- 8: mostly consistent, minor defects
- 6: mixed quality
- 3: major inconsistency
- 1: severe failure

Required output:
Return one JSON object only:
{
  "overall_score": integer 1-10,
  "dimension_scores": {
    "semantic_consistency": integer 1-10,
    "subject_attribute_consistency": integer 1-10,
    "spatial_consistency": integer 1-10,
    "style_lighting_consistency": integer 1-10,
    "detail_integrity": integer 1-10
  },
  "confidence": number 0-1,
  "reasons": ["short reason 1", "short reason 2"],
  "failure_tags": ["optional_tag"]
}
No markdown and no extra text."""


@dataclass
class Sample:
    game_id: str
    theme_item_id: Any
    segment_index: int
    sample_id: str
    image_path: Path
    prompt_text: str
    scene_text: str
    prev_image_path: Optional[Path]
    prev_scene_text: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Score image consistency per game and export Excel.")
    parser.add_argument(
        "--experiment-root",
        type=Path,
        default=DEFAULT_EXPERIMENT_ROOT,
        help="Path to DN-experiment-2.0 root.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help="Directory for JSONL and XLSX outputs.",
    )
    parser.add_argument(
        "--models",
        type=str,
        default=(os.getenv("COHERENCE_MODELS") or "").strip(),
        help="Comma-separated judge models. Falls back to env or VISION_REF_MODEL.",
    )
    parser.add_argument(
        "--max-samples",
        type=int,
        default=0,
        help="Optional limit of scored samples, 0 means no limit.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Build sample table and Excel skeleton without model calls.",
    )
    return parser.parse_args()


def load_env() -> None:
    if load_dotenv is not None:
        load_dotenv(REPO_ROOT / ".env")


def parse_models(models_csv: str) -> List[str]:
    names = [m.strip() for m in (models_csv or "").split(",") if m.strip()]
    if names:
        return names
    fallback = (os.getenv("VISION_REF_MODEL") or os.getenv("COHERENCE_MODEL_A") or "").strip()
    return [fallback] if fallback else []


def env_str(name: str) -> str:
    return (os.getenv(name) or "").strip()


def _data_url_from_image(path: Path) -> str:
    mime = "image/png"
    if path.suffix.lower() in {".jpg", ".jpeg"}:
        mime = "image/jpeg"
    raw = path.read_bytes()
    b64 = base64.b64encode(raw).decode("ascii")
    return f"data:{mime};base64,{b64}"


def _safe_text(value: Any, max_len: int = 1200) -> str:
    if value is None:
        return ""
    text = str(value)
    text = re.sub(r"\s+", " ", text).strip()
    return text[:max_len]


def _extract_json(raw: str) -> Optional[Dict[str, Any]]:
    if not raw:
        return None
    cleaned = raw.strip()
    if "```" in cleaned:
        cleaned = re.sub(r"^```(?:json)?", "", cleaned).replace("```", "").strip()
    start = cleaned.find("{")
    end = cleaned.rfind("}")
    if start < 0 or end < 0 or end <= start:
        return None
    try:
        return json.loads(cleaned[start : end + 1])
    except json.JSONDecodeError:
        return None


def _to_float(v: Any, fallback: float) -> float:
    try:
        return float(v)
    except Exception:
        return fallback


def _normalize_scores(obj: Dict[str, Any]) -> Dict[str, Any]:
    ds = obj.get("dimension_scores") or {}
    out_ds: Dict[str, float] = {}
    for d in DIMENSIONS:
        out_ds[d] = max(SCORE_MIN, min(SCORE_MAX, round(_to_float(ds.get(d), SCORE_DEFAULT))))
    overall = _to_float(obj.get("overall_score"), mean(out_ds.values()))
    overall = max(SCORE_MIN, min(SCORE_MAX, round(overall)))
    confidence = _to_float(obj.get("confidence"), 0.5)
    confidence = max(0.0, min(1.0, confidence))
    reasons = obj.get("reasons") or []
    if not isinstance(reasons, list):
        reasons = [str(reasons)]
    reasons = [str(x)[:180] for x in reasons if str(x).strip()]
    if not reasons:
        reasons = ["No reason provided by judge."]
    failure_tags = obj.get("failure_tags") or []
    if not isinstance(failure_tags, list):
        failure_tags = [str(failure_tags)]
    failure_tags = [str(x)[:60] for x in failure_tags if str(x).strip()]
    return {
        "overall_score": overall,
        "dimension_scores": out_ds,
        "confidence": confidence,
        "reasons": reasons,
        "failure_tags": failure_tags,
    }


def build_samples(experiment_root: Path) -> List[Sample]:
    samples: List[Sample] = []
    manifest_files = sorted(experiment_root.glob("theme_*/*_image_paths.json"))
    for manifest_path in manifest_files:
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        game_id = str(manifest.get("game_id") or manifest_path.parent.name)
        theme_item_id = manifest.get("theme_item_id")
        prev_image: Optional[Path] = None
        prev_scene = ""
        for seg in manifest.get("segments", []):
            if not seg.get("exists"):
                continue
            rel = seg.get("image_path_repo_relative")
            if not rel:
                continue
            seg_idx = int(seg.get("segment_index") or 0)
            json_file = str(seg.get("json_file") or "")
            prompt_text = ""
            scene_text = ""
            if json_file:
                p = manifest_path.parent / json_file
                if p.is_file():
                    try:
                        seg_data = json.loads(p.read_text(encoding="utf-8"))
                        prompt_text = _safe_text(seg_data.get("prompt"), 1400)
                        scene_text = _safe_text(seg_data.get("scene"), 1400)
                    except Exception:
                        pass
            image_path = REPO_ROOT / str(rel).replace("/", os.sep)
            if not image_path.is_file():
                continue
            sample_id = f"{game_id}_seg_{seg_idx:03d}"
            samples.append(
                Sample(
                    game_id=game_id,
                    theme_item_id=theme_item_id,
                    segment_index=seg_idx,
                    sample_id=sample_id,
                    image_path=image_path,
                    prompt_text=prompt_text,
                    scene_text=scene_text,
                    prev_image_path=prev_image,
                    prev_scene_text=prev_scene,
                )
            )
            prev_image = image_path
            prev_scene = scene_text
    samples.sort(key=lambda s: (s.game_id, s.segment_index))
    return samples


def score_sample(client: OpenAI, model: str, sample: Sample) -> Dict[str, Any]:
    user_text = (
        f"game_id: {sample.game_id}\n"
        f"segment_index: {sample.segment_index}\n"
        f"Current scene text: {sample.scene_text or '(none)'}\n"
        f"Current prompt text: {sample.prompt_text or '(none)'}\n"
        f"Previous scene text: {sample.prev_scene_text or '(none)'}\n"
        "Evaluate consistency of the current image in the game sequence."
    )
    content: List[Dict[str, Any]] = [{"type": "text", "text": user_text}]
    if sample.prev_image_path is not None and sample.prev_image_path.is_file():
        content.append({"type": "text", "text": "Previous image:"})
        content.append({"type": "image_url", "image_url": {"url": _data_url_from_image(sample.prev_image_path)}})
    content.append({"type": "text", "text": "Current image:"})
    content.append({"type": "image_url", "image_url": {"url": _data_url_from_image(sample.image_path)}})

    resp = client.chat.completions.create(
        model=model,
        temperature=0.1,
        max_tokens=900,
        messages=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": content},
        ],
    )
    raw = (resp.choices[0].message.content or "").strip()
    parsed = _extract_json(raw) or {}
    normalized = _normalize_scores(parsed)
    normalized["raw_response"] = raw[:4000]
    return normalized


def aggregate(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    bucket: Dict[Tuple[str, str], List[Dict[str, Any]]] = defaultdict(list)
    for r in rows:
        bucket[(r["game_id"], r["judge_model"])].append(r)
    out: List[Dict[str, Any]] = []
    for (game_id, model), vals in sorted(bucket.items()):
        out.append(
            {
                "game_id": game_id,
                "judge_model": model,
                "sample_count": len(vals),
                "overall_score_mean": round(mean([v["overall_score"] for v in vals]), 4),
                "semantic_consistency_mean": round(mean([v["semantic_consistency"] for v in vals]), 4),
                "subject_attribute_consistency_mean": round(mean([v["subject_attribute_consistency"] for v in vals]), 4),
                "spatial_consistency_mean": round(mean([v["spatial_consistency"] for v in vals]), 4),
                "style_lighting_consistency_mean": round(mean([v["style_lighting_consistency"] for v in vals]), 4),
                "detail_integrity_mean": round(mean([v["detail_integrity"] for v in vals]), 4),
                "confidence_mean": round(mean([v["confidence"] for v in vals]), 4),
            }
        )
    return out


def aggregate_ensemble(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    bucket: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for r in rows:
        bucket[r["game_id"]].append(r)
    out: List[Dict[str, Any]] = []
    for game_id, vals in sorted(bucket.items()):
        models = sorted({str(v.get("judge_model", "")) for v in vals if v.get("judge_model")})
        out.append(
            {
                "game_id": game_id,
                "model_count": len(models),
                "models": ",".join(models),
                "row_count": len(vals),
                "overall_score_mean": round(mean([v["overall_score"] for v in vals]), 4),
                "semantic_consistency_mean": round(mean([v["semantic_consistency"] for v in vals]), 4),
                "subject_attribute_consistency_mean": round(mean([v["subject_attribute_consistency"] for v in vals]), 4),
                "spatial_consistency_mean": round(mean([v["spatial_consistency"] for v in vals]), 4),
                "style_lighting_consistency_mean": round(mean([v["style_lighting_consistency"] for v in vals]), 4),
                "detail_integrity_mean": round(mean([v["detail_integrity"] for v in vals]), 4),
                "confidence_mean": round(mean([v["confidence"] for v in vals]), 4),
            }
        )
    return out


def write_excel(
    path: Path,
    raw_rows: List[Dict[str, Any]],
    summary_rows: List[Dict[str, Any]],
    ensemble_rows: List[Dict[str, Any]],
) -> None:
    if Workbook is None:
        raise RuntimeError("openpyxl is not installed; cannot write xlsx. Re-run with --dry-run or install openpyxl.")
    wb = Workbook()
    ws_raw = wb.active
    ws_raw.title = "per_image_scores"
    raw_header = [
        "game_id",
        "theme_item_id",
        "segment_index",
        "sample_id",
        "judge_model",
        "overall_score",
        "semantic_consistency",
        "subject_attribute_consistency",
        "spatial_consistency",
        "style_lighting_consistency",
        "detail_integrity",
        "confidence",
        "reasons",
        "failure_tags",
        "image_path",
    ]
    ws_raw.append(raw_header)
    for r in raw_rows:
        ws_raw.append([r.get(k, "") for k in raw_header])

    ws_sum = wb.create_sheet("per_game_summary")
    sum_header = [
        "game_id",
        "judge_model",
        "sample_count",
        "overall_score_mean",
        "semantic_consistency_mean",
        "subject_attribute_consistency_mean",
        "spatial_consistency_mean",
        "style_lighting_consistency_mean",
        "detail_integrity_mean",
        "confidence_mean",
    ]
    ws_sum.append(sum_header)
    for r in summary_rows:
        ws_sum.append([r.get(k, "") for k in sum_header])

    ws_ens = wb.create_sheet("per_game_ensemble_mean")
    ens_header = [
        "game_id",
        "model_count",
        "models",
        "row_count",
        "overall_score_mean",
        "semantic_consistency_mean",
        "subject_attribute_consistency_mean",
        "spatial_consistency_mean",
        "style_lighting_consistency_mean",
        "detail_integrity_mean",
        "confidence_mean",
    ]
    ws_ens.append(ens_header)
    for r in ensemble_rows:
        ws_ens.append([r.get(k, "") for k in ens_header])

    wb.save(path)


def jsonl_write(path: Path, rows: Iterable[Dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8") as f:
        for r in rows:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")


def main() -> int:
    load_env()
    args = parse_args()
    models = parse_models(args.models)
    if not models and not args.dry_run:
        print("No judge model configured. Set --models or COHERENCE_MODELS.", file=sys.stderr)
        return 2
    api_key = env_str("COHERENCE_API_KEY") or env_str("VISION_REF_API_KEY") or env_str("Origin_Segment_Analyst_API_KEY")
    base_url = env_str("COHERENCE_BASE_URL") or env_str("VISION_REF_BASE_URL") or env_str("Origin_Segment_Analyst_BASE_URL") or "https://api.openai.com/v1"
    if not args.dry_run and not api_key:
        print("Missing API key in environment.", file=sys.stderr)
        return 2

    samples = build_samples(args.experiment_root)
    if args.max_samples > 0:
        samples = samples[: args.max_samples]

    args.output_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    raw_jsonl = args.output_dir / f"per_game_image_scores_{ts}.jsonl"
    latest_jsonl = args.output_dir / "latest_per_game_image_scores.jsonl"
    excel_path = args.output_dir / f"per_game_image_consistency_{ts}.xlsx"
    latest_excel = args.output_dir / "latest_per_game_image_consistency.xlsx"
    summary_json = args.output_dir / f"per_game_image_consistency_summary_{ts}.json"

    raw_rows: List[Dict[str, Any]] = []
    if args.dry_run:
        for s in samples:
            raw_rows.append(
                {
                    "game_id": s.game_id,
                    "theme_item_id": s.theme_item_id,
                    "segment_index": s.segment_index,
                    "sample_id": s.sample_id,
                    "judge_model": "dry-run",
                    "overall_score": "",
                    "semantic_consistency": "",
                    "subject_attribute_consistency": "",
                    "spatial_consistency": "",
                    "style_lighting_consistency": "",
                    "detail_integrity": "",
                    "confidence": "",
                    "reasons": "",
                    "failure_tags": "",
                    "image_path": str(s.image_path),
                }
            )
    else:
        client = OpenAI(api_key=api_key, base_url=base_url)
        for idx, s in enumerate(samples, start=1):
            for model in models:
                result = score_sample(client, model, s)
                raw_rows.append(
                    {
                        "game_id": s.game_id,
                        "theme_item_id": s.theme_item_id,
                        "segment_index": s.segment_index,
                        "sample_id": s.sample_id,
                        "judge_model": model,
                        "overall_score": result["overall_score"],
                        "semantic_consistency": result["dimension_scores"]["semantic_consistency"],
                        "subject_attribute_consistency": result["dimension_scores"]["subject_attribute_consistency"],
                        "spatial_consistency": result["dimension_scores"]["spatial_consistency"],
                        "style_lighting_consistency": result["dimension_scores"]["style_lighting_consistency"],
                        "detail_integrity": result["dimension_scores"]["detail_integrity"],
                        "confidence": result["confidence"],
                        "reasons": " | ".join(result["reasons"]),
                        "failure_tags": ",".join(result["failure_tags"]),
                        "image_path": str(s.image_path),
                    }
                )
            if idx % 5 == 0:
                print(f"scored {idx}/{len(samples)} samples")

    scored_rows = [r for r in raw_rows if isinstance(r.get("overall_score"), (int, float))]
    summary_rows = aggregate(scored_rows)
    ensemble_rows = aggregate_ensemble(scored_rows)
    jsonl_write(raw_jsonl, raw_rows)
    jsonl_write(latest_jsonl, raw_rows)
    write_excel(excel_path, raw_rows, summary_rows, ensemble_rows)
    write_excel(latest_excel, raw_rows, summary_rows, ensemble_rows)

    payload = {
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "experiment_root": str(args.experiment_root),
        "dry_run": args.dry_run,
        "models": models if not args.dry_run else ["dry-run"],
        "samples": len(samples),
        "rows": len(raw_rows),
        "output_jsonl": str(raw_jsonl),
        "latest_jsonl": str(latest_jsonl),
        "output_excel": str(excel_path),
        "latest_excel": str(latest_excel),
    }
    summary_json.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
