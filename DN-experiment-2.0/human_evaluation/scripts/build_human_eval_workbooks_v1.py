#!/usr/bin/env python3
"""Build pilot human-evaluation workbooks for DN text and image consistency."""
from __future__ import annotations

import csv
import importlib.util
import json
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

try:
    from PIL import Image as PILImage
except Exception:  # pragma: no cover
    PILImage = None

ROOT = Path(r"C:\Users\User\Desktop\DN-main")
EXP_ROOT = ROOT / "DN-experiment-2.0"
HUMAN_ROOT = EXP_ROOT / "human_evaluation"
OUT_DIR = HUMAN_ROOT / "workbooks"
TEXT_XLSX = OUT_DIR / "human_text_consistency_rating_v1.xlsx"
IMAGE_XLSX = OUT_DIR / "human_image_consistency_rating_v1.xlsx"
NOTES_MD = OUT_DIR / "human_eval_dataset_build_notes_v1.md"
MANIFEST_CSV = OUT_DIR / "human_eval_manifest_v1.csv"
CHECK_CSV = OUT_DIR / "human_eval_rating_fields_check.csv"
CHECK_JSON = OUT_DIR / "human_eval_rating_fields_check_summary.json"
THUMB_DIR = OUT_DIR / "image_thumbnails_v1"

TEXT_N = 20
IMAGE_N = 50
MAX_STORY_CHARS = 5500
SKIP_IMAGE_GAME_IDS = {"game_1776418475_uho8y4"}  # User requested replacing old IMG_0015-IMG_0024.

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUB_FILL = PatternFill("solid", fgColor="D9EAF7")
TECH_FILL = PatternFill("solid", fgColor="E7E6E6")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TEXT_TAGS = "identity_conflict, goal_drift, causal_gap, time_order_error, scene_jump, unresolved_thread, incoherent_story"
IMAGE_TAGS = "scene_mismatch, subject_drift, count_mismatch, missing_key_object, wrong_action, spatial_incoherent, lighting_conflict, style_drift, artifact_heavy, text_garbled"
SCORE_LIST = '"1,2,3,4,5"'
CONF_LIST = '"1.0,0.9,0.8,0.7,0.6,0.5,0.4"'
BIN_LIST = '"0,1"'

TECH_LEAK_PATTERNS = [
    re.compile(p, re.I) for p in [
        r"claude", r"gemini", r"gpt-4", r"gpt-5", r"sonnet", r"judge_model", r"api[_ -]?key", r"overall_score\s*[:=]",
        r"semantic_consistency\s*[:=]", r"llm", r"config_name", r"no_council", r"pregen", r"default_20",
    ]
]


def read_json(path: Path) -> dict[str, Any] | None:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def clean_text(text: Any) -> str:
    if text is None:
        return ""
    s = str(text).replace("\r\n", "\n").replace("\r", "\n")
    s = re.sub(r"[ \t]+", " ", s)
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s.strip()


def truncate(text: str, limit: int) -> str:
    text = clean_text(text)
    if len(text) <= limit:
        return text
    return text[: limit - 30].rstrip() + "\n\n……（为便于人工评分，已截断过长内容）"


def theme_dirs() -> list[Path]:
    return sorted([p for p in EXP_ROOT.glob("theme_*") if p.is_dir()], key=lambda p: p.name)


def display_summary(story: str) -> str:
    paras = [p.strip() for p in story.split("\n") if p.strip()]
    if not paras:
        return "请阅读剧情文本，判断主线是否连贯。"
    first = paras[0]
    second = paras[1] if len(paras) > 1 else ""
    base = first if len(first) >= 60 else (first + " " + second).strip()
    return truncate(base, 220)


def collect_text_samples() -> list[dict[str, Any]]:
    samples: list[dict[str, Any]] = []
    for d in theme_dirs():
        jsons = sorted([p for p in d.glob("*.json") if not p.name.endswith("manifest.json") and "image_paths" not in p.name])
        for start in (0, 3):
            window = jsons[start : start + 3]
            scenes = []
            meta = None
            for p in window:
                data = read_json(p) or {}
                scene = clean_text(data.get("scene", ""))
                if scene:
                    scenes.append(f"【片段 {data.get('segment_index', len(scenes)+1)}】\n{scene}")
                if meta is None and data:
                    meta = data
            if not scenes or meta is None:
                continue
            story = truncate("\n\n".join(scenes), MAX_STORY_CHARS)
            idx = len(samples) + 1
            seg_label = f"{start + 1}-{start + len(window)}"
            samples.append({
                "sample_id": f"TEXT_{idx:04d}",
                "display_id": f"TEXT_{idx:04d}",
                "modality": "text",
                "game_id": meta.get("game_id", ""),
                "theme_item_id": meta.get("theme_item_id", ""),
                "segment_index": seg_label,
                "source_path": str(d),
                "text_context_path": ";".join(str(p) for p in window),
                "image_path": "",
                "summary": display_summary(story),
                "story": story,
                "hint": "请只根据剧情文本判断：角色身份、核心目标、时间顺序、场景转换和因果链是否连贯。非 5 分必须写具体证据。",
                "calibration": "否",
            })
            if len(samples) >= TEXT_N:
                break
        if len(samples) >= TEXT_N:
            break
    if samples:
        samples[0]["calibration"] = "是"
    return samples


def load_image_manifest() -> list[dict[str, Any]]:
    manifest = EXP_ROOT / "图片一致性_experiment" / "multiview_image_consistency" / "results" / "latest_eval_manifest.jsonl"
    rows = []
    if not manifest.exists():
        return rows
    for line in manifest.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        try:
            rows.append(json.loads(line))
        except json.JSONDecodeError:
            pass
    return rows


def candidate_json_for_image(row: dict[str, Any], image_path: Path) -> Path | None:
    direct = image_path.with_suffix(".json")
    if direct.exists():
        return direct
    gid = row.get("game_id", "")
    seg = int(row.get("segment_index") or 0)
    theme_item_id = row.get("theme_item_id", "")
    for d in theme_dirs():
        if gid and gid not in d.name:
            continue
        for p in d.glob("*.json"):
            if p.name.endswith("manifest.json") or "image_paths" in p.name:
                continue
            data = read_json(p) or {}
            if str(data.get("segment_index", "")) == str(seg) and str(data.get("theme_item_id", "")) == str(theme_item_id):
                return p
    return None


def list_join(v: Any, limit: int = 5) -> str:
    if isinstance(v, list):
        vals = [clean_text(x) for x in v if clean_text(x)]
        return "；".join(vals[:limit])
    if isinstance(v, dict):
        parts = []
        for key, val in v.items():
            part = list_join(val, limit=2)
            if part:
                parts.append(part)
        return "；".join(parts[:limit])
    return clean_text(v)


def cn_key_people(data: dict[str, Any], scene: str) -> str:
    names = []
    for n in ["林守诚", "张叔", "小翠", "账房老李", "厨子王大"]:
        if n in scene and n not in names:
            names.append(n)
    pj = data.get("prompt_json") or {}
    chars = list_join(((pj.get("environment") or {}).get("characters") or []), 3)
    base = "、".join(names) if names else "主角及当前段落中出现的关键人物"
    if chars:
        return base + "。参考要求：" + chars
    return base


def cn_key_objects(scene: str, prompt: str) -> str:
    candidates = ["铜制印信", "玉门驿", "驿站大门", "油灯", "账册", "马", "枯叶", "木门", "包裹", "信件", "刀", "火把"]
    found = [x for x in candidates if x in scene or x in prompt]
    return "、".join(found[:6]) if found else "当前段落中推动剧情的关键物体；若画面缺失核心物体，请降分。"


def cn_scene_action(data: dict[str, Any], scene: str) -> str:
    pj = data.get("prompt_json") or {}
    env = list_join((pj.get("environment") or {}).get("background") or [], 3)
    comp = list_join(pj.get("composition") or [], 3)
    first = display_summary(scene)
    parts = [first]
    if env:
        parts.append("场景要求：" + env)
    if comp:
        parts.append("动作/构图要求：" + comp)
    return "\n".join(parts)


def collect_image_samples() -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    candidates: list[dict[str, Any]] = []
    exclusions: list[dict[str, Any]] = []
    for row in load_image_manifest():
        if row.get("game_id") in SKIP_IMAGE_GAME_IDS:
            exclusions.append({**row, "image_path": "", "exclusion_reason": "user_requested_replace_current_IMG_0015_to_IMG_0024"})
            continue
        rel = row.get("image_path_repo_relative", "")
        image_path = ROOT / rel if rel else Path(row.get("image_path", ""))
        if not image_path.exists():
            exclusions.append({**row, "image_path": str(image_path), "exclusion_reason": "image_path_missing"})
            continue
        json_path = candidate_json_for_image(row, image_path)
        if not json_path or not json_path.exists():
            exclusions.append({**row, "image_path": str(image_path), "exclusion_reason": "context_json_missing"})
            continue
        data = read_json(json_path) or {}
        scene = clean_text(data.get("scene", ""))
        if not scene:
            exclusions.append({**row, "image_path": str(image_path), "exclusion_reason": "scene_text_missing"})
            continue
        prompt = clean_text(data.get("prompt", ""))
        candidates.append({
            "sample_id": "",
            "display_id": "",
            "modality": "image",
            "game_id": row.get("game_id", data.get("game_id", "")),
            "theme_item_id": row.get("theme_item_id", data.get("theme_item_id", "")),
            "segment_index": row.get("segment_index", data.get("segment_index", "")),
            "image_path": str(image_path),
            "source_path": str(json_path),
            "text_context_path": str(json_path),
            "scene_summary": display_summary(scene),
            "story_text": truncate(scene, 1200),
            "key_people": cn_key_people(data, scene),
            "key_objects": cn_key_objects(scene, prompt),
            "key_scene_action": cn_scene_action(data, scene),
            "previous_context": "如无特别说明，只评价本行图片与当前场景/剧情段落是否一致。",
            "calibration": "否",
        })
    grouped: dict[str, list[dict[str, Any]]] = {}
    for item in candidates:
        grouped.setdefault(str(item["game_id"]), []).append(item)
    # Human image consistency needs local continuity: keep the same game together,
    # and sort each game's frames by segment_index.
    ordered: list[dict[str, Any]] = []
    for key in sorted(grouped):
        group = sorted(grouped[key], key=lambda x: int(x.get("segment_index") or 0))
        ordered.extend(group)
    samples = ordered[:IMAGE_N]
    for idx, item in enumerate(samples, start=1):
        item["sample_id"] = f"IMG_{idx:04d}"
        item["display_id"] = f"IMG_{idx:04d}"
    if samples:
        samples[0]["calibration"] = "是"
    return samples, exclusions


def style_sheet(ws, header_row: int = 1, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    for cell in ws[header_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER


def add_validation(ws, col: int, start: int, end: int, formula: str, prompt: str = "") -> None:
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    if prompt:
        dv.prompt = prompt
        dv.showInputMessage = True
    ws.add_data_validation(dv)
    dv.add(f"{get_column_letter(col)}{start}:{get_column_letter(col)}{end}")


def set_widths(ws, widths: dict[int, float]) -> None:
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width


def add_instruction_sheet(wb: Workbook, title: str, lines: list[list[Any]]) -> None:
    ws = wb.active
    ws.title = title
    for r, row in enumerate(lines, start=1):
        for c, value in enumerate(row, start=1):
            ws.cell(r, c, value)
    ws.merge_cells("A1:D1")
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = HEADER_FILL
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
    set_widths(ws, {1: 24, 2: 34, 3: 44, 4: 44})
    for i in range(1, ws.max_row + 1):
        ws.row_dimensions[i].height = 34 if i > 1 else 28
    ws.freeze_panes = "A2"


def build_text_workbook(samples: list[dict[str, Any]]) -> None:
    wb = Workbook()
    lines = [
        ["文本一致性人工评分说明", "", "", ""],
        ["本表用途", "评估剧情文本是否连贯。", "评测员不需要懂模型或代码，只需要按阅读感受和文本证据评分。", ""],
        ["5 分", "完美连贯。", "角色行为一致，场景切换自然，情节有因果关系，整体有完整主线。", ""],
        ["4 分", "较好连贯。", "有轻微逻辑瑕疵，但不影响整体理解。", ""],
        ["3 分", "基本连贯。", "有明显情节跳跃，但能勉强看出主线。", ""],
        ["2 分", "较差连贯。", "角色/场景频繁突变，逻辑混乱，难以理解。", ""],
        ["1 分", "完全不连贯。", "句子之间毫无关联，无法形成故事。", ""],
        ["严格规则", "身份/目标/时间/因果矛盾未解释，不能给 5 分。", "需要大量脑补关键因果，最高给 4 分；主线可懂但跳跃明显，通常给 3 分；中途像换故事，给 1 或 2 分。", "非 5 分必须填写具体理由。"],
        ["confidence", "1.0：非常确定", "0.8：比较确定；0.6：有一定不确定；0.4 或更低：证据不足或很难判断。", ""],
        ["问题标签提示", TEXT_TAGS, "可多选，用英文逗号分隔。", ""],
    ]
    add_instruction_sheet(wb, "评分说明", lines)

    ws = wb.create_sheet("待评文本样本")
    headers = ["样本编号", "主题/背景摘要", "剧情文本", "评测提示", "是否为校准样本", "sample_id", "game_id", "theme_item_id", "source_path", "config_name_or_group", "text_context_path"]
    ws.append(headers)
    for s in samples:
        ws.append([s["display_id"], s["summary"], s["story"], s["hint"], s["calibration"], s["sample_id"], s["game_id"], s["theme_item_id"], s["source_path"], "", s["text_context_path"]])
    style_sheet(ws)
    set_widths(ws, {1: 14, 2: 36, 3: 86, 4: 36, 5: 14, 6: 14, 7: 26, 8: 14, 9: 46, 10: 20, 11: 46})
    for col in range(6, 12):
        ws.column_dimensions[get_column_letter(col)].hidden = True
    for i in range(2, ws.max_row + 1):
        ws.row_dimensions[i].height = 180
    for cell in ws[1][5:]:
        cell.fill = TECH_FILL

    ws = wb.create_sheet("文本评分表")
    headers = ["样本编号", "评分员编号", "剧情连贯性评分（1-5）", "评分信心（0-1）", "是否存在严重硬伤（0/1）", "评分理由", "问题标签", "备注", "填写提示", "sample_id", "modality", "rater_id", "text_coherence_1to5", "confidence", "disqualifying_defect_0or1", "reasons", "failure_tags", "comment"]
    ws.append(headers)
    for i, s in enumerate(samples, start=2):
        ws.cell(i, 1, s["display_id"])
        ws.cell(i, 9, "非 5 分请写具体理由；问题标签可从说明页复制。")
        ws.cell(i, 10, f"=A{i}")
        ws.cell(i, 11, "text")
        ws.cell(i, 12, f"=B{i}")
        ws.cell(i, 13, f"=C{i}")
        ws.cell(i, 14, f"=D{i}")
        ws.cell(i, 15, f"=E{i}")
        ws.cell(i, 16, f"=F{i}")
        ws.cell(i, 17, f"=G{i}")
        ws.cell(i, 18, f"=H{i}")
    style_sheet(ws)
    set_widths(ws, {1: 14, 2: 14, 3: 18, 4: 16, 5: 20, 6: 48, 7: 34, 8: 28, 9: 34, 10: 14, 11: 12, 12: 14, 13: 20, 14: 14, 15: 22, 16: 40, 17: 34, 18: 28})
    for c in range(10, 19):
        ws.column_dimensions[get_column_letter(c)].hidden = True
        ws.cell(1, c).fill = TECH_FILL
    for c in [3]:
        add_validation(ws, c, 2, max(200, len(samples)+20), SCORE_LIST, "请选择 1-5 分")
    add_validation(ws, 4, 2, max(200, len(samples)+20), CONF_LIST, "请选择评分信心")
    add_validation(ws, 5, 2, max(200, len(samples)+20), BIN_LIST, "0=否，1=是")
    for c in [3,4,5,6,7,8]:
        for r in range(2, ws.max_row + 1):
            ws.cell(r, c).fill = INPUT_FILL
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 52
    wb.save(TEXT_XLSX)


def build_image_workbook(samples: list[dict[str, Any]]) -> None:
    wb = Workbook()
    lines = [
        ["图片一致性人工评分说明", "", "", ""],
        ["本表用途", "评估图片是否和剧情/场景要求一致。", "评测员只需要看图片和文字说明，不需要懂模型或代码。", ""],
        ["通用 5 分", "强一致，无实质缺陷。", "", ""],
        ["通用 4 分", "大体一致，有轻微缺陷但不影响理解。", "", ""],
        ["通用 3 分", "质量混合，有明显不一致或不确定性。", "", ""],
        ["通用 2 分", "重大不一致，会影响理解。", "", ""],
        ["通用 1 分", "严重失败，与意图不符或图像结构崩坏。", "", ""],
        ["总体一致性", "综合判断图片是否符合场景要求。", "", ""],
        ["语义一致性", "图片有没有画出剧情核心事件。", "", ""],
        ["主体属性一致性", "人物身份、数量、服装、关键物体是否正确稳定。", "", ""],
        ["空间一致性", "位置关系、比例、透视、场景结构是否合理。", "", ""],
        ["风格与光照一致性", "画风、色调、光照方向和强度是否统一。", "", ""],
        ["细节完整性", "手、脸、局部结构、乱码、水印、伪影等是否有明显问题。", "", ""],
        ["严格规则", "关键人物缺失、身份错误或关键事件没画出来，总体一致性最高 3 分。", "图片很好看但不符合剧情，语义一致性必须低分；主角变脸、服装漂移、人物数量错误，主体属性应降到 2-3 分。", "总体一致性通常不应高于第二低维度分数 + 1；非 5 分必须填写具体理由。"],
        ["问题标签提示", IMAGE_TAGS, "可多选，用英文逗号分隔。", ""],
    ]
    add_instruction_sheet(wb, "评分说明", lines)

    ws = wb.create_sheet("待评图片样本")
    headers = ["样本编号", "图片", "高清原图路径（可复制打开）", "场景说明", "剧情文本/当前段落", "需要关注的关键人物", "需要关注的关键物体", "需要关注的场景/动作", "前序参考说明", "是否为校准样本", "sample_id", "game_id", "theme_item_id", "segment_index", "image_path", "source_path", "config_name_or_group", "text_context_path"]
    ws.append(headers)
    for row_idx, s in enumerate(samples, start=2):
        ws.append([s["display_id"], "", s["image_path"], s["scene_summary"], s["story_text"], s["key_people"], s["key_objects"], s["key_scene_action"], s["previous_context"], s["calibration"], s["sample_id"], s["game_id"], s["theme_item_id"], s["segment_index"], s["image_path"], s["source_path"], "", s["text_context_path"]])
        ws.cell(row_idx, 3).hyperlink = s["image_path"]
        ws.cell(row_idx, 3).style = "Hyperlink"
        img_path = Path(s["image_path"])
        if img_path.exists():
            try:
                # Embed lightweight thumbnails so the workbook stays easy to share.
                embed_path = img_path
                if PILImage is not None:
                    THUMB_DIR.mkdir(parents=True, exist_ok=True)
                    embed_path = THUMB_DIR / f"{s['sample_id']}.jpg"
                    with PILImage.open(img_path) as im:
                        im = im.convert("RGB")
                        im.thumbnail((520, 360))
                        im.save(embed_path, "JPEG", quality=95, optimize=True)
                img = XLImage(str(embed_path))
                img.width = 420
                img.height = 290
                ws.add_image(img, f"B{row_idx}")
            except Exception:
                ws.cell(row_idx, 2, str(img_path))
    style_sheet(ws)
    set_widths(ws, {1: 12, 2: 58, 3: 48, 4: 34, 5: 58, 6: 36, 7: 28, 8: 46, 9: 30, 10: 14, 11: 12, 12: 26, 13: 14, 14: 14, 15: 58, 16: 58, 17: 20, 18: 58})
    for col in range(11, 19):
        ws.column_dimensions[get_column_letter(col)].hidden = True
    for cell in ws[1][10:]:
        cell.fill = TECH_FILL
    for i in range(2, ws.max_row + 1):
        ws.row_dimensions[i].height = 220

    ws = wb.create_sheet("图片评分表")
    headers = ["样本编号", "评分员编号", "总体一致性（1-5）", "语义一致性（1-5）", "主体属性一致性（1-5）", "空间一致性（1-5）", "风格与光照一致性（1-5）", "细节完整性（1-5）", "评分信心（0-1）", "是否存在严重硬伤（0/1）", "评分理由", "问题标签", "备注", "填写提示", "sample_id", "modality", "rater_id", "overall_score", "semantic_consistency", "subject_attribute_consistency", "spatial_consistency", "style_lighting_consistency", "detail_integrity", "confidence", "disqualifying_defect_0or1", "reasons", "failure_tags", "comment"]
    ws.append(headers)
    for i, s in enumerate(samples, start=2):
        ws.cell(i, 1, s["display_id"])
        ws.cell(i, 14, "非 5 分请写具体理由；总体分通常不高于第二低维度分数 + 1。")
        ws.cell(i, 15, f"=A{i}")
        ws.cell(i, 16, "image")
        ws.cell(i, 17, f"=B{i}")
        for src_col, dst_col in [(3,18),(4,19),(5,20),(6,21),(7,22),(8,23),(9,24),(10,25),(11,26),(12,27),(13,28)]:
            ws.cell(i, dst_col, f"={get_column_letter(src_col)}{i}")
    style_sheet(ws)
    widths = {1:12,2:14,3:16,4:16,5:20,6:16,7:22,8:16,9:16,10:20,11:50,12:34,13:28,14:38,
              15:12,16:12,17:14,18:14,19:22,20:28,21:20,22:26,23:18,24:14,25:22,26:40,27:34,28:28}
    set_widths(ws, widths)
    for c in range(15, 29):
        ws.column_dimensions[get_column_letter(c)].hidden = True
        ws.cell(1, c).fill = TECH_FILL
    for c in [3,4,5,6,7,8]:
        add_validation(ws, c, 2, max(300, len(samples)+20), SCORE_LIST, "请选择 1-5 分")
    add_validation(ws, 9, 2, max(300, len(samples)+20), CONF_LIST, "请选择评分信心")
    add_validation(ws, 10, 2, max(300, len(samples)+20), BIN_LIST, "0=否，1=是")
    for c in [3,4,5,6,7,8,9,10,11,12,13]:
        for r in range(2, ws.max_row + 1):
            ws.cell(r, c).fill = INPUT_FILL
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 52
    wb.save(IMAGE_XLSX)


def write_manifest(text_samples: list[dict[str, Any]], image_samples: list[dict[str, Any]], exclusions: list[dict[str, Any]]) -> None:
    fields = ["sample_id", "modality", "display_id", "game_id", "theme_item_id", "segment_index", "source_path", "image_path", "text_context_path", "included_0or1", "exclusion_reason"]
    with MANIFEST_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for s in text_samples + image_samples:
            writer.writerow({k: s.get(k, "") for k in fields} | {"included_0or1": 1, "exclusion_reason": ""})
        for ex in exclusions:
            writer.writerow({
                "sample_id": ex.get("sample_id", ""),
                "modality": "image",
                "display_id": "",
                "game_id": ex.get("game_id", ""),
                "theme_item_id": ex.get("theme_item_id", ""),
                "segment_index": ex.get("segment_index", ""),
                "source_path": ex.get("source_manifest", ""),
                "image_path": ex.get("image_path", ""),
                "text_context_path": "",
                "included_0or1": 0,
                "exclusion_reason": ex.get("exclusion_reason", ""),
            })


def write_notes(text_samples: list[dict[str, Any]], image_samples: list[dict[str, Any]], exclusions: list[dict[str, Any]], check_results: dict[str, Any]) -> None:
    source_lines = [
        f"- `{EXP_ROOT / 'theme_*'}`：用于抽取连续剧情文本、图片路径和图片对应场景说明。",
        f"- `{EXP_ROOT / '图片一致性_experiment' / 'multiview_image_consistency' / 'results' / 'latest_eval_manifest.jsonl'}`：用于定位可用图片样本清单。",
        f"- `{HUMAN_ROOT / 'templates' / 'human_rating_template_v1.csv'}` 和 `{HUMAN_ROOT / 'schemas' / 'human_rating_v1.schema.json'}`：用于复用英文标准评分字段。",
        f"- `{HUMAN_ROOT / 'scripts' / 'summarize_human_ratings.py'}`：用于字段可汇总性轻量检查。",
    ]
    exclusion_counts: dict[str, int] = {}
    for ex in exclusions:
        reason = ex.get("exclusion_reason", "unknown")
        exclusion_counts[reason] = exclusion_counts.get(reason, 0) + 1
    notes = [
        "# Human Evaluation Dataset Build Notes v1",
        "",
        "## 数据源",
        *source_lines,
        "",
        "## 样本数量",
        f"- 文本样本：{len(text_samples)} 条（pilot）。",
        f"- 图片样本：{len(image_samples)} 条（pilot）。",
        f"- 校准样本：文本 1 条、图片 1 条，已在样本页标记。",
        "",
        "## 样本选择与整理",
        "- 文本样本从每个 theme 目录中抽取前 3 个连续剧情片段，合并为可阅读的中文剧情文本。",
        "- 图片样本从图片一致性 manifest 中抽取，逐条确认本地图片存在，并读取相邻 JSON 中的中文剧情段落作为场景说明。",
        "- 所有正式样本使用 `TEXT_0001` / `IMG_0001` 形式匿名化；评测员不需要看到原始 game id 或实验路径。",
        "- 给评测员看的内容包括：样本编号、自然语言摘要/剧情/图片、关注人物/物体/场景要求、评分说明。",
        "- 不给评测员看的内容包括：模型名、实验组名、LLM 分数、API 信息、生成参数、JSON 原始结构。技术追踪列已隐藏或放在右侧。",
        "",
        "## 排除的数据类型",
        "- 图片路径不存在的样本。",
        "- 缺少可读剧情/场景 JSON 上下文的图片样本。",
        "- 只有 LLM 评分但无法组成自然语言测评材料的记录。",
        "- 任何正式样本均未纳入 LLM judge 分数或模型名称。",
        "",
        "## 失效图片路径",
        f"- 排除记录总数：{len(exclusions)}。",
        "- 按原因统计：" + (", ".join(f"{k}={v}" for k, v in sorted(exclusion_counts.items())) if exclusion_counts else "无。"),
        "",
        "## 匿名化和字段设计",
        "- `sample_id` 与展示编号一致，均为匿名 ID。",
        "- 评分表右侧保留英文标准字段：文本 `text_coherence_1to5`；图片 `overall_score`、`semantic_consistency`、`subject_attribute_consistency`、`spatial_consistency`、`style_lighting_consistency`、`detail_integrity` 等。",
        "- 英文字段通过公式引用中文填写列，便于后续另存为 CSV 后使用现有汇总脚本。",
        "",
        "## 质量检查结果",
        f"- 文本 Excel 存在：{TEXT_XLSX.exists()}。",
        f"- 图片 Excel 存在：{IMAGE_XLSX.exists()}。",
        f"- 文本评分字段检查：{check_results.get('text_score_field_ok')}",
        f"- 图片评分字段检查：{check_results.get('image_score_fields_ok')}",
        f"- 图片路径存在性检查：{check_results.get('all_image_paths_exist')}",
        f"- 技术信息泄露扫描：{check_results.get('no_tech_leaks_in_visible_cells')}",
        f"- 现有汇总脚本读取检查：{check_results.get('summarizer_check_ok')}",
        "",
        "## 不完整或需人工确认",
        "- 图片关键人物/物体/动作说明基于当前 JSON 的剧情文本和 prompt_json 自动整理，建议项目负责人抽查 3-5 条确认说明是否符合实验意图。",
        "- 本版本是 pilot 包，适合先发给 2-3 名评测员独立评分；如需 full 版本，可复用构建脚本扩大抽样数量。",
    ]
    NOTES_MD.write_text("\n".join(notes), encoding="utf-8")


def visible_cell_values(path: Path, sheet_names: list[str]) -> list[str]:
    wb = load_workbook(path, data_only=False)
    vals = []
    for name in sheet_names:
        ws = wb[name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                col_letter = get_column_letter(cell.column)
                if ws.column_dimensions[col_letter].hidden:
                    continue
                vals.append(str(cell.value))
    return vals


def validate(text_samples: list[dict[str, Any]], image_samples: list[dict[str, Any]]) -> dict[str, Any]:
    results: dict[str, Any] = {}
    for p in [TEXT_XLSX, IMAGE_XLSX]:
        results[f"{p.name}_exists"] = p.exists()
        wb = load_workbook(p, read_only=False, data_only=False)
        results[f"{p.name}_sheet_count"] = len(wb.sheetnames)
    text_wb = load_workbook(TEXT_XLSX, data_only=False)
    image_wb = load_workbook(IMAGE_XLSX, data_only=False)
    text_headers = [c.value for c in text_wb["文本评分表"][1]]
    image_headers = [c.value for c in image_wb["图片评分表"][1]]
    results["text_score_field_ok"] = "text_coherence_1to5" in text_headers
    required_img = ["overall_score", "semantic_consistency", "subject_attribute_consistency", "spatial_consistency", "style_lighting_consistency", "detail_integrity"]
    results["image_score_fields_ok"] = all(h in image_headers for h in required_img)
    results["all_image_paths_exist"] = all(Path(s["image_path"]).exists() for s in image_samples)
    visible = "\n".join(visible_cell_values(TEXT_XLSX, ["评分说明", "待评文本样本", "文本评分表"]) + visible_cell_values(IMAGE_XLSX, ["评分说明", "待评图片样本", "图片评分表"]))
    leaks = [pat.pattern for pat in TECH_LEAK_PATTERNS if pat.search(visible)]
    results["no_tech_leaks_in_visible_cells"] = not leaks
    results["tech_leak_patterns"] = leaks

    # Create a tiny standards-compliant CSV and run the existing summarizer to verify downstream fields are readable.
    fields = ["sample_id","modality","rater_id","game_id","theme_item_id","segment_index","config_name","source_path","text_context_path","image_path","text_coherence_1to5","overall_score","semantic_consistency","subject_attribute_consistency","spatial_consistency","style_lighting_consistency","detail_integrity","confidence","disqualifying_defect_0or1","reasons","failure_tags","adjudication_needed_0or1","adjudicated_score","comment"]
    with CHECK_CSV.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerow({"sample_id":"TEXT_0001","modality":"text","rater_id":"check","text_coherence_1to5":"4","confidence":"0.8","disqualifying_defect_0or1":"0","reasons":"field check","failure_tags":""})
        writer.writerow({"sample_id":"IMG_0001","modality":"image","rater_id":"check","overall_score":"4","semantic_consistency":"4","subject_attribute_consistency":"4","spatial_consistency":"4","style_lighting_consistency":"4","detail_integrity":"4","confidence":"0.8","disqualifying_defect_0or1":"0","reasons":"field check","failure_tags":""})
    summarizer = HUMAN_ROOT / "scripts" / "summarize_human_ratings.py"
    spec = importlib.util.spec_from_file_location("summarize_human_ratings", summarizer)
    mod = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(mod)
    with CHECK_CSV.open("r", encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    summary, _ = mod.summarize(rows, 4.0)
    CHECK_JSON.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    results["summarizer_check_ok"] = summary.get("row_count") == 2 and summary.get("sample_count") == 2
    return results


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    text_samples = collect_text_samples()
    image_samples, exclusions = collect_image_samples()
    if not text_samples:
        raise RuntimeError("No text samples found")
    if not image_samples:
        raise RuntimeError("No image samples found")
    build_text_workbook(text_samples)
    build_image_workbook(image_samples)
    write_manifest(text_samples, image_samples, exclusions)
    check_results = validate(text_samples, image_samples)
    write_notes(text_samples, image_samples, exclusions, check_results)
    print(json.dumps({
        "text_xlsx": str(TEXT_XLSX),
        "image_xlsx": str(IMAGE_XLSX),
        "notes": str(NOTES_MD),
        "manifest": str(MANIFEST_CSV),
        "text_samples": len(text_samples),
        "image_samples": len(image_samples),
        "checks": check_results,
    }, ensure_ascii=False, indent=2))
    if not all([
        check_results.get("human_text_consistency_rating_v1.xlsx_exists"),
        check_results.get("human_image_consistency_rating_v1.xlsx_exists"),
        check_results.get("text_score_field_ok"),
        check_results.get("image_score_fields_ok"),
        check_results.get("all_image_paths_exist"),
        check_results.get("no_tech_leaks_in_visible_cells"),
        check_results.get("summarizer_check_ok"),
    ]):
        raise SystemExit(2)


if __name__ == "__main__":
    main()
