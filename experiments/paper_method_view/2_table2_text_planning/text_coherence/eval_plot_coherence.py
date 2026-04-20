# -*- coding: utf-8 -*-
"""
DN-experiment-2.0 剧情连贯性 LLM 评估

- 按 manifest 拼接每局全部 segment 的 scene 为完整故事；
- 每个模型对每个样本独立打分 `--runs` 次（默认 3），取算术平均；
- 皮尔逊相关系数：
  - 若配置 ≥2 个可用模型：报告模型两两之间「样本平均分的向量」的 Pearson r；
  - 同时报告每个模型内部各次打分之间的 r（衡量同一模型重复打分稳定性）；
- 结果输出 Excel（openpyxl）。

依赖：openai, tenacity, python-dotenv, openpyxl（见仓库 requirements / requirements-eval.txt）

用法示例：
  # .env 中设置 COHERENCE_MODELS=claude-sonnet-4-20250514,gemini-2.5-flash,gpt-4o
  python DN-experiment-2.0/eval_plot_coherence.py

  # 只评前 2 局、输出到指定文件
  python DN-experiment-2.0/eval_plot_coherence.py --max-games 2 --output coherence.xlsx
"""
from __future__ import annotations

import argparse
import json
import logging
import math
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

_REPO_ROOT = Path(__file__).resolve().parents[1]
_EXPERIMENT_DIR = Path(__file__).resolve().parent

if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

if sys.platform == "win32":
    os.environ.setdefault("PYTHONIOENCODING", "utf-8")
    os.environ.setdefault("PYTHONUTF8", "1")
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

_LOG_DIR = _EXPERIMENT_DIR / "eval_results"
_LOG_DIR.mkdir(parents=True, exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    handlers=[
        logging.FileHandler(_LOG_DIR / "eval_plot_coherence.log", encoding="utf-8", mode="a"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)

try:
    from dotenv import load_dotenv
except ImportError:
    load_dotenv = None  # type: ignore

if load_dotenv is not None:
    load_dotenv(_REPO_ROOT / ".env")

from openai import OpenAI  # noqa: E402
from tenacity import (  # noqa: E402
    retry,
    retry_if_exception_type,
    stop_after_attempt,
    wait_exponential,
)

# ---------------------------------------------------------------------------
# 模型配置
# 优先读 COHERENCE_MODELS（逗号分隔多个模型名），共享同一 API_KEY / BASE_URL。
# 格式与 COUNCIL_MODELS 一致，便于在 .env 中统一管理。
# 若未设置 COHERENCE_MODELS，则回退到旧的 A/B 两模型配置。
# ---------------------------------------------------------------------------
def _default_model_configs() -> List[Dict[str, str]]:
    models_csv = (os.getenv("COHERENCE_MODELS") or "").strip()
    api_key = os.getenv("COHERENCE_API_KEY") or os.getenv("Origin_Segment_Analyst_API_KEY") or ""
    base_url = os.getenv("COHERENCE_BASE_URL") or os.getenv("Origin_Segment_Analyst_BASE_URL") or "https://yunwu.ai/v1"

    if models_csv:
        names = [m.strip() for m in models_csv.split(",") if m.strip()]
        return [
            {"name": n, "model": n, "api_key": api_key, "base_url": base_url}
            for n in names
        ]

    return [
        {
            "name": os.getenv("COHERENCE_MODEL_A_NAME", "annotator-a"),
            "model": os.getenv(
                "COHERENCE_MODEL_A",
                os.getenv("Origin_Segment_Analyst_MODEL", "claude-haiku-4-5-20251001-thinking"),
            ),
            "api_key": os.getenv("Origin_Segment_Analyst_API_KEY", ""),
            "base_url": os.getenv("Origin_Segment_Analyst_BASE_URL", "https://yunwu.ai/v1"),
        },
        {
            "name": os.getenv("COHERENCE_MODEL_B_NAME", "annotator-b"),
            "model": os.getenv(
                "COHERENCE_MODEL_B",
                os.getenv("VISION_REF_MODEL", "gemini-3.1-pro-preview"),
            ),
            "api_key": os.getenv("VISION_REF_API_KEY", ""),
            "base_url": os.getenv("VISION_REF_BASE_URL", "https://yunwu.ai/v1"),
        },
    ]


SYSTEM_PROMPT = """你是一位专业的文学评论家，请严格按照以下标准评估这段故事的剧情连贯性：

剧情连贯性评分标准（1-5分）
5分：完美连贯。角色行为一致，场景切换自然，情节有因果关系，整体有完整主线
4分：较好连贯。有轻微逻辑瑕疵，但不影响整体理解
3分：基本连贯。有明显情节跳跃，但能勉强看出主线
2分：较差连贯。角色/场景频繁突变，逻辑混乱，难以理解
1分：完全不连贯。句子之间毫无关联，无法形成故事

你必须只输出一个 JSON 对象，不要输出其它文字。格式如下：
{"score": <1到5的整数>, "reason": "<一句简短中文理由>"}"""


USER_TEMPLATE = """请评估以下完整故事的剧情连贯性（1-5分）。

## 元信息
- game_id: {game_id}
- 主题: {theme}
- 段数: {segment_count}

## 故事全文

{story_text}
"""


def _build_client(cfg: Dict[str, str]) -> OpenAI:
    return OpenAI(api_key=cfg["api_key"], base_url=cfg["base_url"])


@retry(
    stop=stop_after_attempt(5),
    wait=wait_exponential(multiplier=2, min=3, max=60),
    retry=retry_if_exception_type((ConnectionError, TimeoutError, OSError)),
    reraise=True,
)
def _call_llm(cfg: Dict[str, str], user: str, *, temperature: float) -> str:
    client = _build_client(cfg)
    resp = client.chat.completions.create(
        model=cfg["model"],
        messages=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": user},
        ],
        temperature=temperature,
        max_tokens=2048,
    )
    return (resp.choices[0].message.content or "").strip()


def _strip_thinking(raw: str) -> str:
    cleaned = re.sub(r"<thinking>.*?</thinking>", "", raw, flags=re.DOTALL)
    cleaned = re.sub(r"<reflection>.*?</reflection>", "", cleaned, flags=re.DOTALL)
    return cleaned.strip()


def _extract_json_block(raw: str) -> Optional[str]:
    cleaned = _strip_thinking(raw)
    fence_match = re.search(r"```(?:json)?\s*(\{.*)", cleaned, re.DOTALL)
    if fence_match:
        text = fence_match.group(1)
        end_fence = text.find("```")
        if end_fence != -1:
            text = text[:end_fence]
    else:
        first_brace = cleaned.find("{")
        if first_brace == -1:
            return None
        text = cleaned[first_brace:]

    depth, start_idx = 0, None
    for i, ch in enumerate(text):
        if ch == "{":
            if depth == 0:
                start_idx = i
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0 and start_idx is not None:
                return text[start_idx : i + 1]
    return None


def _parse_score(raw: str) -> Optional[Tuple[int, str]]:
    block = _extract_json_block(raw)
    if not block:
        return None
    try:
        obj = json.loads(block)
    except json.JSONDecodeError:
        return None
    if not isinstance(obj, dict):
        return None
    sc = obj.get("score")
    if isinstance(sc, float) and sc.is_integer():
        sc = int(sc)
    if not isinstance(sc, int) or sc < 1 or sc > 5:
        return None
    reason = obj.get("reason", "")
    if not isinstance(reason, str):
        reason = str(reason)
    return sc, reason


def pearson_r(xs: Sequence[float], ys: Sequence[float]) -> Optional[float]:
    """Pearson 相关系数；长度不一致或方差为 0 时返回 None。"""
    n = len(xs)
    if n != len(ys) or n < 2:
        return None
    mx = sum(xs) / n
    my = sum(ys) / n
    vx = sum((x - mx) ** 2 for x in xs)
    vy = sum((y - my) ** 2 for y in ys)
    if vx <= 0 or vy <= 0:
        return None
    cov = sum((xs[i] - mx) * (ys[i] - my) for i in range(n))
    return cov / math.sqrt(vx * vy)


def discover_games(dataset_dir: Path) -> List[Dict[str, Any]]:
    """扫描 theme_* 目录，优先读 manifest 拼接各段 scene。"""
    games: List[Dict[str, Any]] = []
    for folder in sorted(dataset_dir.iterdir()):
        if not folder.is_dir() or not folder.name.startswith("theme_"):
            continue
        manifests = sorted(folder.glob("*_manifest.json"))
        if not manifests:
            continue
        man_path = manifests[0]
        try:
            man = json.loads(man_path.read_text(encoding="utf-8"))
        except Exception:
            continue
        game_id = man.get("game_id") or folder.name
        theme = (man.get("theme") or "").strip()
        segs = man.get("segments") or []
        if not isinstance(segs, list):
            continue
        ordered = sorted(
            [s for s in segs if isinstance(s, dict)],
            key=lambda x: int(x.get("index", 0)),
        )
        parts: List[str] = []
        for s in ordered:
            jname = s.get("json")
            if not jname:
                continue
            jp = folder / str(jname)
            if not jp.is_file():
                continue
            try:
                seg_data = json.loads(jp.read_text(encoding="utf-8"))
            except Exception:
                continue
            scene = (seg_data.get("scene") or "").strip()
            if scene:
                parts.append(scene)
        if not parts:
            continue
        full_story = "\n\n---\n\n".join(parts)
        first_seg_path = folder / ordered[0].get("json", "")
        theme_item_id = None
        if first_seg_path.is_file():
            try:
                fd = json.loads(first_seg_path.read_text(encoding="utf-8"))
                theme_item_id = fd.get("theme_item_id")
            except Exception:
                pass
        games.append(
            {
                "game_id": game_id,
                "folder": folder.name,
                "theme": theme,
                "theme_item_id": theme_item_id,
                "segment_count": len(parts),
                "full_story": full_story,
            }
        )
    return games


def score_one_game(
    cfg: Dict[str, str],
    game: Dict[str, Any],
    *,
    runs: int,
    temperature: float,
) -> Tuple[Optional[List[int]], List[str], List[str]]:
    """同一模型对单局打 `runs` 次分。返回 (scores 或 None, raw_texts, errors)。"""
    user = USER_TEMPLATE.format(
        game_id=game["game_id"],
        theme=game["theme"] or "(无)",
        segment_count=game["segment_count"],
        story_text=game["full_story"],
    )
    scores: List[int] = []
    raws: List[str] = []
    errs: List[str] = []
    for r in range(runs):
        try:
            raw = _call_llm(cfg, user, temperature=temperature)
            raws.append(raw)
            parsed = _parse_score(raw)
            if parsed is None:
                errs.append(f"第{r + 1}次解析失败")
                scores.append(-1)
            else:
                scores.append(parsed[0])
        except Exception as e:
            errs.append(f"第{r + 1}次调用失败: {e}")
            raws.append("")
            scores.append(-1)
    if any(s < 0 for s in scores):
        return None, raws, errs
    return scores, raws, errs


def write_excel(
    path: Path,
    rows: List[Dict[str, Any]],
    inter_model: List[Tuple[str, str, Optional[float]]],
    intra_by_model: Dict[str, List[Tuple[str, str, Optional[float]]]],
) -> None:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws0 = wb.active
    ws0.title = "scores"

    if not rows:
        ws0.cell(1, 1, "无数据")
    else:
        base_keys = ["game_id", "folder", "theme", "theme_item_id", "segment_count", "story_chars"]
        all_keys: set[str] = set()
        for row in rows:
            all_keys.update(row.keys())
        rest = sorted(k for k in all_keys if k not in base_keys)
        header_ordered = [k for k in base_keys if k in all_keys] + rest

        for c, h in enumerate(header_ordered, 1):
            ws0.cell(1, c, h)
        for r_i, row in enumerate(rows, 2):
            for c, h in enumerate(header_ordered, 1):
                v = row.get(h)
                ws0.cell(r_i, c, v)

        for col in range(1, len(header_ordered) + 1):
            col_letter = get_column_letter(col)
            ws0.column_dimensions[col_letter].width = 16 if col <= 3 else 12

    ws1 = wb.create_sheet("correlation_models")
    ws1.cell(1, 1, "model_a")
    ws1.cell(1, 2, "model_b")
    ws1.cell(1, 3, "pearson_r")
    for i, (a, b, pr) in enumerate(inter_model, 2):
        ws1.cell(i, 1, a)
        ws1.cell(i, 2, b)
        ws1.cell(i, 3, pr if pr is not None else "N/A")

    ws2 = wb.create_sheet("correlation_runs")
    ws2.cell(1, 1, "model")
    ws2.cell(1, 2, "run_pair")
    ws2.cell(1, 3, "pearson_r")
    r2 = 2
    for model_name, pairs in sorted(intra_by_model.items()):
        for a, b, pr in pairs:
            ws2.cell(r2, 1, model_name)
            ws2.cell(r2, 2, f"{a}_vs_{b}")
            ws2.cell(r2, 3, pr if pr is not None else "N/A")
            r2 += 1

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _flatten_row(
    game: Dict[str, Any],
    model_results: Dict[str, Dict[str, Any]],
) -> Dict[str, Any]:
    story = game["full_story"]
    row: Dict[str, Any] = {
        "game_id": game["game_id"],
        "folder": game["folder"],
        "theme": game["theme"],
        "theme_item_id": game["theme_item_id"],
        "segment_count": game["segment_count"],
        "story_chars": len(story),
    }
    for mn, data in model_results.items():
        runs: List[int] = data["scores"]
        for i, s in enumerate(runs, 1):
            row[f"{mn}_run{i}"] = s
        row[f"{mn}_mean"] = round(sum(runs) / len(runs), 4) if runs else ""
        reasons = data.get("reasons") or []
        row[f"{mn}_reason_last"] = reasons[-1] if reasons else ""
    return row


def compute_intra_model_correlations(
    games: List[Dict[str, Any]],
    model_name: str,
    run_count: int,
) -> List[Tuple[str, str, Optional[float]]]:
    """同一模型：各次打分在样本维度上的 Pearson。"""
    pairs: List[Tuple[str, str, Optional[float]]] = []
    for i in range(1, run_count + 1):
        for j in range(i + 1, run_count + 1):
            xs: List[float] = []
            ys: List[float] = []
            for g in games:
                mr = g.get("model_results") or {}
                mdata = mr.get(model_name)
                if not mdata:
                    continue
                sc: List[int] = mdata["scores"]
                if len(sc) >= j:
                    xs.append(float(sc[i - 1]))
                    ys.append(float(sc[j - 1]))
            pairs.append((f"run{i}", f"run{j}", pearson_r(xs, ys)))
    return pairs


def compute_inter_model_correlations(
    games: List[Dict[str, Any]],
    model_names: List[str],
) -> List[Tuple[str, str, Optional[float]]]:
    """模型两两之间：用各样本的平均分向量。"""
    out: List[Tuple[str, str, Optional[float]]] = []
    for i, ma in enumerate(model_names):
        for mb in model_names[i + 1 :]:
            xs: List[float] = []
            ys: List[float] = []
            for g in games:
                mr = g.get("model_results") or {}
                a = mr.get(ma)
                b = mr.get(mb)
                if not a or not b:
                    continue
                sa = a["scores"]
                sb = b["scores"]
                if not sa or not sb:
                    continue
                xs.append(sum(sa) / len(sa))
                ys.append(sum(sb) / len(sb))
            out.append((ma, mb, pearson_r(xs, ys)))
    return out


def main() -> int:
    ap = argparse.ArgumentParser(description="DN-experiment-2.0 剧情连贯性 LLM 评估")
    ap.add_argument(
        "--dataset",
        type=Path,
        default=_EXPERIMENT_DIR,
        help="数据集根目录（含 theme_* 子目录），默认本脚本所在 DN-experiment-2.0",
    )
    ap.add_argument("--runs", type=int, default=3, help="每样本每模型重复打分次数，默认 3")
    ap.add_argument("--temperature", type=float, default=0.4, help="采样温度，默认 0.4")
    ap.add_argument("--max-games", type=int, default=0, help="仅评前 N 局，0 表示全部")
    ap.add_argument(
        "--output",
        type=Path,
        default=None,
        help="输出 xlsx 路径；默认 eval_results/coherence_<timestamp>.xlsx",
    )
    ap.add_argument(
        "--only-first-model",
        action="store_true",
        help="仅使用第一个配置了 api_key 的模型（不做模型间相关）",
    )
    args = ap.parse_args()

    dataset_dir = args.dataset.resolve()
    if not dataset_dir.is_dir():
        log.error("数据集目录不存在: %s", dataset_dir)
        return 2

    configs = _default_model_configs()
    active = [c for c in configs if (c.get("api_key") or "").strip()]
    if args.only_first_model:
        active = active[:1]
    if not active:
        log.error(
            "未找到可用的 API Key。请在 .env 中配置：\n"
            "  COHERENCE_MODELS=model1,model2,model3  （逗号分隔多个模型名）\n"
            "  COHERENCE_API_KEY=你的密钥\n"
            "  COHERENCE_BASE_URL=https://yunwu.ai/v1",
        )
        return 1

    for c in active:
        log.info("启用模型: %s (%s)", c["name"], c["model"])

    games = discover_games(dataset_dir)
    if not games:
        log.error("未在 %s 下发现任何带 manifest 的有效游戏数据", dataset_dir)
        return 1

    if args.max_games and args.max_games > 0:
        games = games[: args.max_games]

    log.info("共 %d 局待评 | 每模型每局 %d 次打分", len(games), args.runs)

    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    out_path = args.output
    if out_path is None:
        out_path = _LOG_DIR / f"coherence_{ts}.xlsx"
    else:
        out_path = out_path.resolve()

    evaluated: List[Dict[str, Any]] = []

    for idx, game in enumerate(games):
        log.info("── [%d/%d] %s ──", idx + 1, len(games), game["game_id"])
        model_results: Dict[str, Dict[str, Any]] = {}
        for cfg in active:
            log.info("  → 模型 %s 打分中...", cfg["name"])
            scores, raws, errs = score_one_game(
                cfg, game, runs=args.runs, temperature=args.temperature
            )
            if errs:
                for e in errs:
                    log.warning("    %s", e)
            if scores is None:
                log.error("    模型 %s 本局失败，跳过该模型本局", cfg["name"])
                continue
            last_reason = ""
            for raw in raws:
                p = _parse_score(raw)
                if p:
                    last_reason = p[1]
            model_results[cfg["name"]] = {
                "scores": scores,
                "reasons": [(_parse_score(r) or (0, ""))[1] for r in raws],
                "raw_last": raws[-1] if raws else "",
                "reason_last": last_reason,
            }
            log.info("    %s: %s → 均值 %.3f", cfg["name"], scores, sum(scores) / len(scores))

        if not model_results:
            log.error("本局所有模型失败，跳过")
            continue

        evaluated.append({"game": game, "model_results": model_results})

    rows: List[Dict[str, Any]] = []
    for item in evaluated:
        game = item["game"]
        mr = item["model_results"]
        row = _flatten_row(game, mr)
        rows.append(row)

    model_names = [c["name"] for c in active]
    inter_model = compute_inter_model_correlations(evaluated, model_names)
    intra_by_model: Dict[str, List[Tuple[str, str, Optional[float]]]] = {}
    for mn in model_names:
        intra_by_model[mn] = compute_intra_model_correlations(evaluated, mn, args.runs)

    log.info("======== 皮尔逊相关（模型间，基于各样本平均分）========")
    for a, b, pr in inter_model:
        log.info("  %s vs %s: r = %s", a, b, f"{pr:.4f}" if pr is not None else "N/A")

    log.info("======== 皮尔逊相关（同模型多次打分稳定性）========")
    for mn, pairs in intra_by_model.items():
        for ra, rb, pr in pairs:
            log.info(
                "  [%s] %s vs %s: r = %s",
                mn,
                ra,
                rb,
                f"{pr:.4f}" if pr is not None else "N/A",
            )

    try:
        write_excel(out_path, rows, inter_model, intra_by_model)
        log.info("已写入: %s", out_path)
    except ImportError:
        log.error("未安装 openpyxl，请执行: pip install openpyxl")
        return 1

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
