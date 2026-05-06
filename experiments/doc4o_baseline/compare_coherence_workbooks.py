# -*- coding: utf-8 -*-
"""Compare DN and DOC-4o LLM coherence score workbooks."""
from __future__ import annotations

import argparse
import json
import statistics
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook


def read_rows(path: Path) -> List[Dict[str, Any]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    if "scores" not in wb.sheetnames:
        raise RuntimeError(f"Workbook has no 'scores' sheet: {path}")
    ws = wb["scores"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    out: List[Dict[str, Any]] = []
    for values in rows[1:]:
        row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
        if row.get("game_id"):
            out.append(row)
    return out


def to_float(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def summarize(label: str, rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    mean_cols = sorted({k for row in rows for k in row if k.endswith("_mean")})
    per_model: Dict[str, Dict[str, Any]] = {}
    all_scores: List[float] = []
    for col in mean_cols:
        vals = [v for row in rows if (v := to_float(row.get(col))) is not None]
        if not vals:
            continue
        model_name = col[: -len("_mean")]
        per_model[model_name] = {
            "sample_count": len(vals),
            "mean": round(statistics.mean(vals), 4),
            "median": round(statistics.median(vals), 4),
            "min": min(vals),
            "max": max(vals),
        }
        all_scores.extend(vals)

    by_theme: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        theme_id = row.get("theme_item_id")
        key = str(theme_id) if theme_id not in (None, "") else str(row.get("theme") or row.get("game_id"))
        vals = [v for col in mean_cols if (v := to_float(row.get(col))) is not None]
        if not vals:
            continue
        by_theme[key] = {
            "theme": row.get("theme"),
            "mean": round(statistics.mean(vals), 4),
            "model_mean_columns": mean_cols,
        }

    return {
        "label": label,
        "sample_count": len(rows),
        "model_mean_columns": mean_cols,
        "overall_mean": round(statistics.mean(all_scores), 4) if all_scores else None,
        "overall_median": round(statistics.median(all_scores), 4) if all_scores else None,
        "per_model": per_model,
        "by_theme": by_theme,
    }


def compare(dn: Dict[str, Any], doc: Dict[str, Any]) -> Dict[str, Any]:
    dn_mean = dn.get("overall_mean")
    doc_mean = doc.get("overall_mean")
    delta = None
    relative = None
    if isinstance(dn_mean, (int, float)) and isinstance(doc_mean, (int, float)):
        delta = round(doc_mean - dn_mean, 4)
        relative = round(delta / dn_mean * 100, 2) if dn_mean else None

    common_theme_ids = sorted(set(dn["by_theme"]).intersection(doc["by_theme"]), key=lambda x: int(x) if x.isdigit() else x)
    theme_rows: List[Dict[str, Any]] = []
    for tid in common_theme_ids:
        dn_theme = dn["by_theme"][tid]
        doc_theme = doc["by_theme"][tid]
        theme_rows.append(
            {
                "theme_id": tid,
                "theme": dn_theme.get("theme") or doc_theme.get("theme"),
                "dn_mean": dn_theme["mean"],
                "doc4o_mean": doc_theme["mean"],
                "delta_doc_minus_dn": round(doc_theme["mean"] - dn_theme["mean"], 4),
            }
        )

    return {
        "dn_overall_mean": dn_mean,
        "doc4o_overall_mean": doc_mean,
        "delta_doc_minus_dn": delta,
        "relative_delta_percent": relative,
        "common_theme_count": len(common_theme_ids),
        "per_theme": theme_rows,
    }


def write_markdown(path: Path, dn: Dict[str, Any], doc: Dict[str, Any], comp: Dict[str, Any]) -> None:
    lines = [
        "# DOC-4o vs DN LLM Coherence Comparison",
        "",
        f"- DN samples: {dn['sample_count']}",
        f"- DOC-4o samples: {doc['sample_count']}",
        f"- DN overall mean: {dn.get('overall_mean')}",
        f"- DOC-4o overall mean: {doc.get('overall_mean')}",
        f"- Delta (DOC-4o - DN): {comp.get('delta_doc_minus_dn')}",
        f"- Relative delta: {comp.get('relative_delta_percent')}%",
        "",
        "| Theme ID | Theme | DN mean | DOC-4o mean | Delta |",
        "|---:|---|---:|---:|---:|",
    ]
    for row in comp["per_theme"]:
        theme = str(row.get("theme") or "").replace("|", "\\|")
        lines.append(
            f"| {row['theme_id']} | {theme} | {row['dn_mean']} | {row['doc4o_mean']} | {row['delta_doc_minus_dn']} |"
        )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="Compare DN and DOC-4o coherence workbooks.")
    parser.add_argument("--dn-xlsx", type=Path, required=True)
    parser.add_argument("--doc-xlsx", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, default=None)
    args = parser.parse_args()

    dn_rows = read_rows(args.dn_xlsx)
    doc_rows = read_rows(args.doc_xlsx)
    dn_summary = summarize("DN", dn_rows)
    doc_summary = summarize("DOC-4o", doc_rows)
    comparison = compare(dn_summary, doc_summary)

    out_dir = args.output_dir or args.doc_xlsx.resolve().parent / "comparison"
    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    payload = {
        "created_at_utc": stamp,
        "dn_xlsx": str(args.dn_xlsx.resolve()),
        "doc_xlsx": str(args.doc_xlsx.resolve()),
        "dn": dn_summary,
        "doc4o": doc_summary,
        "comparison": comparison,
    }
    json_path = out_dir / "doc4o_vs_dn_coherence_summary.json"
    md_path = out_dir / "doc4o_vs_dn_coherence_summary.md"
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    write_markdown(md_path, dn_summary, doc_summary, comparison)

    print(f"Wrote {json_path}")
    print(f"Wrote {md_path}")
    print(f"DN overall mean: {dn_summary.get('overall_mean')}")
    print(f"DOC-4o overall mean: {doc_summary.get('overall_mean')}")
    print(f"Delta (DOC-4o - DN): {comparison.get('delta_doc_minus_dn')}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
