# -*- coding: utf-8 -*-
"""Compare DN and baseline coherence workbooks produced by eval_plot_coherence.py."""
from __future__ import annotations

import argparse
import csv
from pathlib import Path
from statistics import mean
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook, load_workbook


def read_scores(path: Path) -> Tuple[List[Dict[str, Any]], List[str]]:
    wb = load_workbook(path, data_only=True)
    ws = wb["scores"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows: List[Dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        row = {str(headers[c - 1]): ws.cell(r, c).value for c in range(1, ws.max_column + 1) if headers[c - 1]}
        if row.get("theme_item_id") is not None:
            rows.append(row)
    mean_cols = [h for h in headers if isinstance(h, str) and h.endswith("_mean")]
    return rows, mean_cols


def row_score(row: Dict[str, Any], mean_cols: List[str]) -> float | None:
    values = []
    for col in mean_cols:
        value = row.get(col)
        if isinstance(value, (int, float)):
            values.append(float(value))
    if not values:
        return None
    return mean(values)


def key_for(row: Dict[str, Any]) -> str:
    theme_id = row.get("theme_item_id")
    if theme_id is not None:
        return str(theme_id)
    return str(row.get("folder") or row.get("game_id"))


def main() -> int:
    parser = argparse.ArgumentParser(description="Compare coherence scores between DN and a baseline workbook.")
    parser.add_argument("--dn-xlsx", type=Path, required=True)
    parser.add_argument("--baseline-xlsx", type=Path, required=True)
    parser.add_argument("--baseline-name", default="baseline")
    parser.add_argument("--output-xlsx", type=Path, required=True)
    parser.add_argument("--output-csv", type=Path, default=None)
    args = parser.parse_args()

    dn_rows, dn_mean_cols = read_scores(args.dn_xlsx)
    base_rows, base_mean_cols = read_scores(args.baseline_xlsx)

    dn_by_key = {key_for(row): row for row in dn_rows}
    base_by_key = {key_for(row): row for row in base_rows}
    keys = sorted(set(dn_by_key) & set(base_by_key), key=lambda x: int(x) if x.isdigit() else x)

    paired: List[Dict[str, Any]] = []
    for key in keys:
        dn_row = dn_by_key[key]
        base_row = base_by_key[key]
        dn_score = row_score(dn_row, dn_mean_cols)
        base_score = row_score(base_row, base_mean_cols)
        if dn_score is None or base_score is None:
            continue
        paired.append(
            {
                "theme_item_id": key,
                "theme": dn_row.get("theme") or base_row.get("theme"),
                "dn_game_id": dn_row.get("game_id"),
                "baseline_game_id": base_row.get("game_id"),
                "dn_score": round(dn_score, 4),
                f"{args.baseline_name}_score": round(base_score, 4),
                "delta_dn_minus_baseline": round(dn_score - base_score, 4),
            }
        )

    dn_values = [float(row["dn_score"]) for row in paired]
    base_key = f"{args.baseline_name}_score"
    base_values = [float(row[base_key]) for row in paired]
    deltas = [float(row["delta_dn_minus_baseline"]) for row in paired]

    summary = {
        "paired_games": len(paired),
        "dn_mean": round(mean(dn_values), 4) if dn_values else "",
        f"{args.baseline_name}_mean": round(mean(base_values), 4) if base_values else "",
        "mean_delta_dn_minus_baseline": round(mean(deltas), 4) if deltas else "",
        "dn_xlsx": str(args.dn_xlsx.resolve()),
        "baseline_xlsx": str(args.baseline_xlsx.resolve()),
    }

    args.output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "summary"
    for r, (k, v) in enumerate(summary.items(), 1):
        ws.cell(r, 1, k)
        ws.cell(r, 2, v)

    ws2 = wb.create_sheet("paired")
    if paired:
        headers = list(paired[0].keys())
        for c, h in enumerate(headers, 1):
            ws2.cell(1, c, h)
        for r, row in enumerate(paired, 2):
            for c, h in enumerate(headers, 1):
                ws2.cell(r, c, row.get(h))
    wb.save(args.output_xlsx)

    if args.output_csv:
        args.output_csv.parent.mkdir(parents=True, exist_ok=True)
        with args.output_csv.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=list(paired[0].keys()) if paired else list(summary.keys()))
            writer.writeheader()
            if paired:
                writer.writerows(paired)
            else:
                writer.writerow(summary)

    print(summary)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
