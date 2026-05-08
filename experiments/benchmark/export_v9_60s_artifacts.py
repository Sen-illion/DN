import csv
import json
from pathlib import Path
from statistics import mean, median
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font
from PIL import Image, ImageDraw, ImageFont

PROJECT_ROOT = Path(__file__).resolve().parents[2]
base = PROJECT_ROOT / "experiments" / "benchmark" / "standard_runs"

OLD_PROJECT_ROOTS = (
    Path("C:/") / "Users" / "zhang" / "Desktop" / "DN",
    Path("C:/") / "Users" / "User" / "Desktop" / "DN-main",
)

def resolve_project_path(value):
    path = Path(value)
    if path.exists():
        return path
    text = str(path)
    for old_root in OLD_PROJECT_ROOTS:
        old_text = str(old_root)
        if text.lower().startswith(old_text.lower()):
            candidate = PROJECT_ROOT / path.relative_to(old_root)
            if candidate.exists():
                return candidate
    return path

summary_path = base / "benchmark_v9_readwait_60s_merged_12v12_summary.json"
payload = json.loads(summary_path.read_text(encoding="utf-8"))

source_files = payload["source_files"]
combined_runs = []
for group, paths in source_files.items():
    for path in paths:
        source_path = resolve_project_path(path)
        sub = json.loads(source_path.read_text(encoding="utf-8"))
        for run in sub["runs"]:
            second = run.get("second_click", {})
            first = run.get("first_click", {})
            worldview = run.get("worldview", {})
            combined_runs.append({
                "group": group,
                "source_file": source_path.name,
                "benchmark_id": run.get("benchmark_id"),
                "theme_id": run.get("theme_id"),
                "read_wait_s": run.get("read_wait_s"),
                "worldview_elapsed_s": worldview.get("elapsed_s"),
                "first_click_elapsed_s": first.get("elapsed_s"),
                "second_click_elapsed_s": second.get("elapsed_s"),
                "second_click_status": second.get("status"),
                "second_click_real_scene": not second.get("is_placeholder", True),
                "second_click_has_image": second.get("has_image"),
                "second_click_hit": second.get("inferred_cache_result"),
                "selected_option_text": second.get("selected_option_text"),
            })

out_summary_csv = base / "benchmark_v9_readwait_60s_merged_12v12_summary_table.csv"
out_paired_csv = base / "benchmark_v9_readwait_60s_merged_12v12_paired_table.csv"
out_raw_csv = base / "benchmark_v9_readwait_60s_merged_12v12_raw_runs.csv"
out_xlsx = base / "benchmark_v9_readwait_60s_merged_12v12_workbook.xlsx"
out_png_1 = base / "benchmark_v9_readwait_60s_latency_comparison.png"
out_png_2 = base / "benchmark_v9_readwait_60s_real_scene_rate.png"

summary_rows = []
for group in ("off", "on"):
    s = payload["summary"][group]
    row = {
        "group": group,
        "sample_size": s["sample_size"],
        "success_count": s["success_count"],
        "real_scene_count": s["real_scene_count"],
        "real_scene_rate": s["real_scene_rate"],
        "likely_hit_rate": s["likely_hit_rate"],
        "second_click_mean_s": s["second_click"]["mean"],
        "second_click_median_s": s["second_click"]["median"],
        "second_click_p95_s": s["second_click"]["p95"],
        "real_scene_mean_s": s["real_scene_second_click"].get("mean"),
        "real_scene_median_s": s["real_scene_second_click"].get("median"),
        "real_scene_p95_s": s["real_scene_second_click"].get("p95"),
    }
    summary_rows.append(row)

paired_rows = payload["paired"]["rows"]

with out_summary_csv.open("w", encoding="utf-8-sig", newline="") as f:
    writer = csv.DictWriter(f, fieldnames=list(summary_rows[0].keys()))
    writer.writeheader()
    writer.writerows(summary_rows)

with out_paired_csv.open("w", encoding="utf-8-sig", newline="") as f:
    writer = csv.DictWriter(f, fieldnames=list(paired_rows[0].keys()))
    writer.writeheader()
    writer.writerows(paired_rows)

with out_raw_csv.open("w", encoding="utf-8-sig", newline="") as f:
    writer = csv.DictWriter(f, fieldnames=list(combined_runs[0].keys()))
    writer.writeheader()
    writer.writerows(combined_runs)

wb = Workbook()
ws1 = wb.active
ws1.title = "summary"
ws2 = wb.create_sheet("paired")
ws3 = wb.create_sheet("raw_runs")
ws4 = wb.create_sheet("chart_data")

for ws, rows in [(ws1, summary_rows), (ws2, paired_rows), (ws3, combined_runs)]:
    headers = list(rows[0].keys())
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    for row in rows:
        ws.append([row.get(h) for h in headers])

ws4.append(["metric", "off", "on"])
ws4.append(["second_click_median_s", payload["summary"]["off"]["second_click"]["median"], payload["summary"]["on"]["second_click"]["median"]])
ws4.append(["second_click_mean_s", payload["summary"]["off"]["second_click"]["mean"], payload["summary"]["on"]["second_click"]["mean"]])
ws4.append(["real_scene_rate", payload["summary"]["off"]["real_scene_rate"], payload["summary"]["on"]["real_scene_rate"]])
for c in ws4[1]:
    c.font = Font(bold=True)

chart1 = BarChart()
chart1.title = "Second click latency"
chart1.y_axis.title = "seconds"
chart1.x_axis.title = "metric"
data = Reference(ws4, min_col=2, max_col=3, min_row=1, max_row=3)
cats = Reference(ws4, min_col=1, min_row=2, max_row=3)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.height = 7
chart1.width = 12
ws1.add_chart(chart1, "N2")

chart2 = BarChart()
chart2.title = "Real scene rate"
chart2.y_axis.title = "rate"
chart2.x_axis.title = "metric"
data2 = Reference(ws4, min_col=2, max_col=3, min_row=1, max_row=4)
cats2 = Reference(ws4, min_col=1, min_row=4, max_row=4)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)
chart2.height = 7
chart2.width = 12
ws1.add_chart(chart2, "N20")

wb.save(out_xlsx)

# simple png charts via pillow
W, H = 900, 520
BG = (250, 248, 243)
TEXT = (35, 31, 32)
OFF = (201, 104, 81)
ON = (71, 122, 198)
GRID = (210, 210, 210)

try:
    font_title = ImageFont.truetype("arial.ttf", 28)
    font_text = ImageFont.truetype("arial.ttf", 20)
    font_small = ImageFont.truetype("arial.ttf", 16)
except Exception:
    font_title = ImageFont.load_default()
    font_text = ImageFont.load_default()
    font_small = ImageFont.load_default()

# chart 1 latency
img = Image.new("RGB", (W, H), BG)
d = ImageDraw.Draw(img)
d.text((40, 30), "DN 60s second-click latency (12v12)", fill=TEXT, font=font_title)
left, bottom, top = 120, 430, 120
bar_w = 120
max_v = max(payload["summary"]["off"]["second_click"]["median"], payload["summary"]["on"]["second_click"]["median"], payload["summary"]["off"]["second_click"]["mean"], payload["summary"]["on"]["second_click"]["mean"]) * 1.4
for i in range(6):
    y = top + i * (bottom - top) / 5
    d.line((left, y, W - 60, y), fill=GRID, width=1)
    val = round(max_v * (1 - i / 5), 3)
    d.text((35, y - 8), str(val), fill=TEXT, font=font_small)
metrics = [
    ("median", payload["summary"]["off"]["second_click"]["median"], payload["summary"]["on"]["second_click"]["median"]),
    ("mean", payload["summary"]["off"]["second_click"]["mean"], payload["summary"]["on"]["second_click"]["mean"]),
]
for idx, (label, off_v, on_v) in enumerate(metrics):
    x0 = left + idx * 280
    for j, (name, val, color) in enumerate([("off", off_v, OFF), ("on", on_v, ON)]):
        x = x0 + j * (bar_w + 30)
        h = (val / max_v) * (bottom - top)
        d.rectangle((x, bottom - h, x + bar_w, bottom), fill=color)
        d.text((x + 30, bottom + 10), name, fill=TEXT, font=font_text)
        d.text((x + 18, bottom - h - 24), f"{val:.3f}", fill=TEXT, font=font_small)
    d.text((x0 + 80, 455), label, fill=TEXT, font=font_text)
img.save(out_png_1)

# chart 2 real scene rate
img = Image.new("RGB", (W, H), BG)
d = ImageDraw.Draw(img)
d.text((40, 30), "DN 60s real-scene hit rate (12v12)", fill=TEXT, font=font_title)
left, bottom, top = 180, 420, 120
max_v = 1.0
for i in range(6):
    y = top + i * (bottom - top) / 5
    d.line((left, y, W - 60, y), fill=GRID, width=1)
    val = round(max_v * (1 - i / 5), 2)
    d.text((105, y - 8), str(val), fill=TEXT, font=font_small)
vals = [("off", payload["summary"]["off"]["real_scene_rate"], OFF), ("on", payload["summary"]["on"]["real_scene_rate"], ON)]
for i, (name, val, color) in enumerate(vals):
    x = left + i * 220
    h = (val / max_v) * (bottom - top)
    d.rectangle((x, bottom - h, x + 140, bottom), fill=color)
    d.text((x + 45, bottom + 12), name, fill=TEXT, font=font_text)
    d.text((x + 42, bottom - h - 24), f"{val:.3f}", fill=TEXT, font=font_small)
img.save(out_png_2)

print(out_summary_csv)
print(out_paired_csv)
print(out_raw_csv)
print(out_xlsx)
print(out_png_1)
print(out_png_2)
